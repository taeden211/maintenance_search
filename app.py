from __future__ import annotations

import json
import os
import re
import sqlite3
import subprocess
import sys
import threading
import traceback
from copy import copy
from collections import Counter
from dataclasses import asdict, dataclass
from pathlib import Path
from queue import Empty, Queue
from typing import Callable, Optional

import joblib
import numpy as np
from openpyxl import load_workbook
from sklearn.feature_extraction.text import TfidfVectorizer

import tkinter as tk
from tkinter import filedialog, messagebox, ttk


TOKEN_RE = re.compile(r"[0-9A-Za-z가-힣]+")

YEAR_RE = re.compile(r"(20\d{2})")
FILE_RE = re.compile(r"(20\d{2})(0[1-9]|1[0-2])")

APP_TITLE = "유지보수 사례 검색기"
CACHE_DIR_NAME = ".maintenance_search_cache"
CACHE_DB_NAME = "maintenance_cases.sqlite"
CACHE_ARTIFACT_NAME = "search_artifacts.joblib"
CACHE_REPORT_NAME = "index_report.json"
DEFAULT_TOP_N = 20
SEARCH_INDEX_VERSION = 5
MANIFEST_CHECK_INTERVAL_MS = 30_000
QUERY_HINT = "장애, 조치, 부서, 사용자, 날짜 등을 입력하세요"
EXCLUDE_FILE_KEYWORDS = (
    "검색결과",
    "보고서",
    "분석",
    "result",
    "output",
    "backup",
    "bak",
)
CHECKED_VALUES = {"O", "Y", "YES", "TRUE", "1", "○", "예"}


@dataclass(slots=True)
class MaintenanceCase:
    # 원본 Excel 행 위치와 주요 값을 함께 들고 있어 검색 결과에서 바로 원본 확인/수정이 가능하다.
    case_id: int
    source_file: str
    source_sheet: str
    row_num: int
    sequence_number: str
    year: int
    month: int
    date_text: str
    department: str
    user: str
    issue_text: str
    action_text: str
    apc: str
    pc_filter: str
    utmp: str
    sheet_title: str
    search_text: str


@dataclass(slots=True)
class SearchFilters:
    year_from: Optional[int] = None
    year_to: Optional[int] = None
    month_from: Optional[tuple[int, int]] = None
    month_to: Optional[tuple[int, int]] = None
    department: str = ""
    user: str = ""
    require_apc: bool = False
    require_pc_filter: bool = False
    require_utmp: bool = False


class BM25Index:
    # 짧은 장애 문장에서는 단어 일치가 직관적이어서 TF-IDF와 함께 BM25 점수를 사용한다.
    def __init__(self, tokenized_docs: list[list[str]], k1: float = 1.5, b: float = 0.75) -> None:
        self.k1 = k1
        self.b = b
        self.tokenized_docs = tokenized_docs
        self.term_freqs = [Counter(doc) for doc in tokenized_docs]
        self.doc_len = np.array([len(doc) for doc in tokenized_docs], dtype=np.float32)
        self.avgdl = float(self.doc_len.mean()) if len(self.doc_len) else 0.0
        self.doc_freq: dict[str, int] = {}
        for doc in tokenized_docs:
            for term in set(doc):
                self.doc_freq[term] = self.doc_freq.get(term, 0) + 1
        self.idf: dict[str, float] = {}
        n_docs = len(tokenized_docs)
        for term, freq in self.doc_freq.items():
            self.idf[term] = float(np.log(1.0 + (n_docs - freq + 0.5) / (freq + 0.5)))

    def score(self, query_tokens: list[str]) -> np.ndarray:
        if not self.tokenized_docs or not query_tokens:
            return np.zeros(len(self.tokenized_docs), dtype=np.float32)

        scores = np.zeros(len(self.tokenized_docs), dtype=np.float32)
        for term in query_tokens:
            if term not in self.idf:
                continue
            idf = self.idf[term]
            for idx, term_freq in enumerate(self.term_freqs):
                tf = term_freq.get(term, 0)
                if tf == 0:
                    continue
                denom = tf + self.k1 * (1.0 - self.b + self.b * (self.doc_len[idx] / self.avgdl if self.avgdl else 0.0))
                scores[idx] += idf * (tf * (self.k1 + 1.0)) / denom
        return scores


@dataclass(slots=True)
class SearchArtifacts:
    records: list[MaintenanceCase]
    vectorizer: TfidfVectorizer
    tfidf_matrix: object
    bm25: BM25Index
    searchable_texts: list[str]
    tokenized_texts: list[list[str]]
    years: list[int]


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\u3000", " ").strip()


def tokenize(text: str) -> list[str]:
    return [token for token in TOKEN_RE.findall(text.lower()) if token.strip()]


def parse_year_month_from_path(path: Path) -> tuple[int, int]:
    for candidate in [path.stem, path.name, path.parent.name]:
        match = FILE_RE.search(candidate)
        if match:
            return int(match.group(1)), int(match.group(2))
        match = YEAR_RE.search(candidate)
        if match:
            return int(match.group(1)), 1
    return 0, 0


def safe_cell(value: object) -> str:
    return normalize_text(value)


def is_checked(value: str) -> bool:
    return normalize_text(value).strip().upper() in CHECKED_VALUES


def get_excel_skip_reason(path: Path) -> str:
    if path.name.startswith("~$"):
        return "엑셀 임시 파일"
    lowered = path.name.lower()
    for keyword in EXCLUDE_FILE_KEYWORDS:
        if keyword.lower() in lowered:
            return f"제외 키워드 포함: {keyword}"
    return ""


def build_source_file_manifest(folder: Path) -> list[dict[str, object]]:
    manifest: list[dict[str, object]] = []
    for path in sorted(folder.rglob("*.xlsx")):
        if get_excel_skip_reason(path):
            continue
        try:
            stat = path.stat()
            relative_path = path.relative_to(folder).as_posix()
        except OSError:
            continue
        manifest.append(
            {
                "path": relative_path,
                "size": int(stat.st_size),
                "mtime_ns": int(stat.st_mtime_ns),
            }
        )
    return manifest


def normalize_manifest_entry(entry: object) -> Optional[tuple[str, int, int]]:
    if not isinstance(entry, dict):
        return None
    path = str(entry.get("path", "")).replace("\\", "/")
    if not path:
        return None
    try:
        size = int(entry.get("size", 0))
        mtime_ns = int(entry.get("mtime_ns", 0))
    except (TypeError, ValueError):
        return None
    return path, size, mtime_ns


def compare_source_file_manifest(folder: Path, report: dict[str, object]) -> tuple[bool, str]:
    saved_manifest = report.get("file_manifest")
    if not isinstance(saved_manifest, list):
        return True, "이전 인덱스 형식입니다."

    saved_entries = [entry for entry in (normalize_manifest_entry(item) for item in saved_manifest) if entry]
    current_entries = [entry for entry in (normalize_manifest_entry(item) for item in build_source_file_manifest(folder)) if entry]
    if sorted(saved_entries) == sorted(current_entries):
        return False, ""

    saved_by_path = {path: (size, mtime_ns) for path, size, mtime_ns in saved_entries}
    current_by_path = {path: (size, mtime_ns) for path, size, mtime_ns in current_entries}
    added = sorted(set(current_by_path) - set(saved_by_path))
    removed = sorted(set(saved_by_path) - set(current_by_path))
    changed = sorted(
        path
        for path in set(saved_by_path) & set(current_by_path)
        if saved_by_path[path] != current_by_path[path]
    )

    parts: list[str] = []
    if added:
        parts.append(f"추가 {len(added)}개")
    if removed:
        parts.append(f"삭제 {len(removed)}개")
    if changed:
        parts.append(f"수정 {len(changed)}개")
    return True, ", ".join(parts) if parts else "엑셀 파일 변경"


def looks_like_maintenance_sheet(sheet: object) -> bool:
    texts: list[str] = []
    for row in sheet.iter_rows(min_row=1, max_row=8, max_col=10, values_only=True):
        for value in row:
            text = safe_cell(value).replace(" ", "")
            if text:
                texts.append(text)

    has_issue = any("장애내용" in text or text == "장애" for text in texts)
    has_action = any("조치내용" in text or "조치" in text for text in texts)
    has_department = any("부서" in text for text in texts)
    has_date = any("날짜" in text for text in texts)
    return has_issue and has_action and (has_department or has_date)


def create_empty_index_report() -> dict[str, object]:
    return {
        "total_files": 0,
        "read_files": 0,
        "excluded_files": 0,
        "skipped_files": 0,
        "total_rows": 0,
        "valid_cases": 0,
        "excluded_rows": 0,
        "multi_sheet_files": 0,
        "missing_departments": 0,
        "missing_users": 0,
        "missing_apc": 0,
        "missing_pc_filter": 0,
        "missing_utmp": 0,
        "excluded_file_details": [],
        "skipped_file_details": [],
        "multi_sheet_details": [],
        "sheet_details": [],
        "excluded_row_details": [],
        "quality_details": [],
        "file_manifest": [],
    }


def add_report_detail(
    report: dict[str, object],
    detail_key: str,
    detail_type: str,
    file_name: str,
    path: str,
    sheet: str,
    row: int,
    reason: str,
    date: str,
    department: str,
    user: str,
    issue: str,
    action: str,
) -> None:
    report[detail_key].append(
        {
            "type": detail_type,
            "file": file_name,
            "path": path,
            "sheet": sheet,
            "row": row,
            "reason": reason,
            "date": date,
            "department": department,
            "user": user,
            "issue": issue,
            "action": action,
        }
    )


class MaintenanceRepository:
    # SQLite와 joblib 캐시는 선택한 데이터 폴더 아래에만 저장해 배포 파일과 원본 데이터를 분리한다.
    def __init__(self, folder: Path) -> None:
        self.folder = folder
        self.cache_dir = folder / CACHE_DIR_NAME
        self.db_path = self.cache_dir / CACHE_DB_NAME
        self.artifact_path = self.cache_dir / CACHE_ARTIFACT_NAME
        self.report_path = self.cache_dir / CACHE_REPORT_NAME

    def ensure_cache_dir(self) -> None:
        self.cache_dir.mkdir(parents=True, exist_ok=True)

    def clear(self) -> None:
        self.ensure_cache_dir()
        if self.db_path.exists():
            self.db_path.unlink()
        if self.artifact_path.exists():
            self.artifact_path.unlink()
        if self.report_path.exists():
            self.report_path.unlink()

    def save_records(self, records: list[MaintenanceCase]) -> None:
        self.ensure_cache_dir()
        with sqlite3.connect(self.db_path) as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS cases (
                    case_id INTEGER PRIMARY KEY,
                    source_file TEXT,
                    source_sheet TEXT,
                    row_num INTEGER,
                    sequence_number TEXT,
                    year INTEGER,
                    month INTEGER,
                    date_text TEXT,
                    department TEXT,
                    user TEXT,
                    issue_text TEXT,
                    action_text TEXT,
                    apc TEXT,
                    pc_filter TEXT,
                    utmp TEXT,
                    sheet_title TEXT,
                    search_text TEXT
                )
                """
            )
            conn.execute("DELETE FROM cases")
            conn.executemany(
                """
                INSERT INTO cases (
                    case_id, source_file, source_sheet, row_num, sequence_number, year, month, date_text,
                    department, user, issue_text, action_text, apc, pc_filter, utmp,
                    sheet_title, search_text
                ) VALUES (
                    :case_id, :source_file, :source_sheet, :row_num, :sequence_number, :year, :month, :date_text,
                    :department, :user, :issue_text, :action_text, :apc, :pc_filter, :utmp,
                    :sheet_title, :search_text
                )
                """,
                [asdict(record) for record in records],
            )
            conn.commit()

    def load_records(self) -> list[MaintenanceCase]:
        if not self.db_path.exists():
            return []
        with sqlite3.connect(self.db_path) as conn:
            conn.row_factory = sqlite3.Row
            rows = conn.execute("SELECT * FROM cases ORDER BY case_id").fetchall()
        return [MaintenanceCase(**dict(row)) for row in rows]

    def save_artifacts(self, artifacts: SearchArtifacts) -> None:
        self.ensure_cache_dir()
        joblib.dump(
            {
                "index_version": SEARCH_INDEX_VERSION,
                "records": artifacts.records,
                "vectorizer": artifacts.vectorizer,
                "tfidf_matrix": artifacts.tfidf_matrix,
                "bm25": artifacts.bm25,
                "searchable_texts": artifacts.searchable_texts,
                "tokenized_texts": artifacts.tokenized_texts,
                "years": artifacts.years,
            },
            self.artifact_path,
        )

    def load_artifacts(self) -> Optional[SearchArtifacts]:
        if not self.artifact_path.exists():
            return None
        payload = joblib.load(self.artifact_path)
        if payload.get("index_version") != SEARCH_INDEX_VERSION:
            return None
        return SearchArtifacts(
            records=payload["records"],
            vectorizer=payload["vectorizer"],
            tfidf_matrix=payload["tfidf_matrix"],
            bm25=payload["bm25"],
            searchable_texts=payload["searchable_texts"],
            tokenized_texts=payload["tokenized_texts"],
            years=payload["years"],
        )

    def save_report(self, report: dict[str, object]) -> None:
        self.ensure_cache_dir()
        self.report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    def load_report(self) -> Optional[dict[str, object]]:
        if not self.report_path.exists():
            return None
        return json.loads(self.report_path.read_text(encoding="utf-8"))

    def compare_source_manifest(self, report: dict[str, object]) -> tuple[bool, str]:
        return compare_source_file_manifest(self.folder, report)


class MaintenanceSearchEngine:
    # 폴더의 Excel 파일을 읽어 검색 인덱스를 만들고, 캐시가 오래되었는지 판단한다.
    def __init__(self) -> None:
        self.repository: Optional[MaintenanceRepository] = None
        self.artifacts: Optional[SearchArtifacts] = None
        self.records: list[MaintenanceCase] = []
        self.loaded_folder: Optional[Path] = None
        self.index_report: Optional[dict[str, object]] = None
        self.index_stale = False
        self.index_stale_message = ""

    @property
    def is_ready(self) -> bool:
        return self.artifacts is not None and bool(self.records)

    def load_folder(self, folder: Path) -> bool:
        self.repository = MaintenanceRepository(folder)
        artifacts = self.repository.load_artifacts()
        if artifacts is None:
            return False
        self.artifacts = artifacts
        self.records = artifacts.records
        self.loaded_folder = folder
        self.index_report = self.repository.load_report() or self._legacy_report(len(self.records))
        self.index_stale, self.index_stale_message = self.repository.compare_source_manifest(self.index_report)
        return True

    def build_from_folder(self, folder: Path) -> tuple[int, list[int], dict[str, object]]:
        repository = MaintenanceRepository(folder)
        repository.clear()
        records, report = self._collect_cases(folder)
        if not records:
            raise RuntimeError("엑셀 파일에서 유효한 장애 사례를 찾지 못했습니다.")

        searchable_texts = [record.search_text for record in records]
        tokenized_texts = [tokenize(text) for text in searchable_texts]

        vectorizer = TfidfVectorizer(
            analyzer="char_wb",
            ngram_range=(2, 5),
            min_df=1,
            sublinear_tf=True,
            norm="l2",
        )
        tfidf_matrix = vectorizer.fit_transform(searchable_texts)
        bm25 = BM25Index(tokenized_texts)
        years = sorted({record.year for record in records if record.year})

        artifacts = SearchArtifacts(
            records=records,
            vectorizer=vectorizer,
            tfidf_matrix=tfidf_matrix,
            bm25=bm25,
            searchable_texts=searchable_texts,
            tokenized_texts=tokenized_texts,
            years=years,
        )
        repository.save_records(records)
        repository.save_artifacts(artifacts)
        report["valid_cases"] = len(records)
        repository.save_report(report)
        self.repository = repository
        self.artifacts = artifacts
        self.records = records
        self.loaded_folder = folder
        self.index_report = report
        self.index_stale = False
        self.index_stale_message = ""
        return len(records), years, report

    def _collect_cases(self, folder: Path) -> tuple[list[MaintenanceCase], dict[str, object]]:
        records: list[MaintenanceCase] = []
        report = create_empty_index_report()
        report["file_manifest"] = build_source_file_manifest(folder)
        case_id = 1

        xlsx_paths = sorted(folder.rglob("*.xlsx"))
        report["total_files"] = len(xlsx_paths)
        for xlsx_path in xlsx_paths:
            relative_path = str(xlsx_path.relative_to(folder))
            skip_reason = get_excel_skip_reason(xlsx_path)
            if skip_reason:
                report["excluded_files"] = int(report["excluded_files"]) + 1
                report["excluded_file_details"].append({"file": xlsx_path.name, "path": relative_path, "reason": skip_reason})
                continue
            year, month = parse_year_month_from_path(xlsx_path)
            try:
                workbook = load_workbook(xlsx_path, data_only=True, read_only=True)
            except Exception as exc:
                report["skipped_files"] = int(report["skipped_files"]) + 1
                report["skipped_file_details"].append({"file": xlsx_path.name, "path": relative_path, "reason": f"엑셀 로딩 실패: {exc}"})
                continue

            try:
                sheet_names = workbook.sheetnames
                read_sheet_count = 0

                for sheet_name in sheet_names:
                    sheet = workbook[sheet_name]
                    if not looks_like_maintenance_sheet(sheet):
                        report["sheet_details"].append(
                            {
                                "file": xlsx_path.name,
                                "path": relative_path,
                                "sheet": sheet_name,
                                "status": "제외",
                                "reason": "유지보수 양식으로 판단되지 않음",
                                "rows": 0,
                                "valid_cases": 0,
                                "excluded_rows": 0,
                            }
                        )
                        continue

                    read_sheet_count += 1
                    sheet_title = safe_cell(sheet.cell(1, 1).value)
                    source_sheet = safe_cell(sheet.title)
                    last_date = ""
                    sheet_rows = 0
                    sheet_valid_cases = 0
                    sheet_excluded_rows = 0

                    for row_num, row_values in enumerate(
                        sheet.iter_rows(min_row=5, max_col=9, values_only=True),
                        start=5,
                    ):
                        sheet_rows += 1
                        report["total_rows"] = int(report["total_rows"]) + 1
                        row_values = tuple(row_values) + (None,) * (9 - len(row_values))

                        sequence_number = safe_cell(row_values[0])
                        date_value = safe_cell(row_values[1])
                        if date_value:
                            last_date = date_value
                        date_text = date_value or last_date
                        department = safe_cell(row_values[2])
                        user = safe_cell(row_values[3])
                        issue_text = safe_cell(row_values[4])
                        action_text = safe_cell(row_values[5])
                        apc = safe_cell(row_values[6])
                        pc_filter = safe_cell(row_values[7])
                        utmp = safe_cell(row_values[8])

                        if not issue_text and not action_text:
                            sheet_excluded_rows += 1
                            report["excluded_rows"] = int(report["excluded_rows"]) + 1
                            add_report_detail(
                                report,
                                "excluded_row_details",
                                "제외 행",
                                xlsx_path.name,
                                relative_path,
                                source_sheet,
                                row_num,
                                "장애내용/조치내용 모두 비어 있음",
                                date_text,
                                department,
                                user,
                                issue_text,
                                action_text,
                            )
                            continue

                        if not department:
                            report["missing_departments"] = int(report["missing_departments"]) + 1
                            add_report_detail(
                                report,
                                "quality_details",
                                "부서 누락",
                                xlsx_path.name,
                                relative_path,
                                source_sheet,
                                row_num,
                                "부서 미입력",
                                date_text,
                                department,
                                user,
                                issue_text,
                                action_text,
                            )
                        if not user:
                            report["missing_users"] = int(report["missing_users"]) + 1
                            add_report_detail(
                                report,
                                "quality_details",
                                "사용자 누락",
                                xlsx_path.name,
                                relative_path,
                                source_sheet,
                                row_num,
                                "사용자 미입력",
                                date_text,
                                department,
                                user,
                                issue_text,
                                action_text,
                            )
                        if not apc:
                            report["missing_apc"] = int(report["missing_apc"]) + 1
                            add_report_detail(
                                report,
                                "quality_details",
                                "APC 누락",
                                xlsx_path.name,
                                relative_path,
                                source_sheet,
                                row_num,
                                "APC 미입력",
                                date_text,
                                department,
                                user,
                                issue_text,
                                action_text,
                            )
                        if not pc_filter:
                            report["missing_pc_filter"] = int(report["missing_pc_filter"]) + 1
                            add_report_detail(
                                report,
                                "quality_details",
                                "PC filter 누락",
                                xlsx_path.name,
                                relative_path,
                                source_sheet,
                                row_num,
                                "PC filter 미입력",
                                date_text,
                                department,
                                user,
                                issue_text,
                                action_text,
                            )
                        if not utmp:
                            report["missing_utmp"] = int(report["missing_utmp"]) + 1
                            add_report_detail(
                                report,
                                "quality_details",
                                "UTMP 누락",
                                xlsx_path.name,
                                relative_path,
                                source_sheet,
                                row_num,
                                "UTMP 미입력",
                                date_text,
                                department,
                                user,
                                issue_text,
                                action_text,
                            )

                        searchable_text = " ".join(
                            part
                            for part in [
                                issue_text,
                                action_text,
                                department,
                                user,
                                date_text,
                                sheet_title,
                                xlsx_path.stem,
                            ]
                            if part
                        )

                        records.append(
                            MaintenanceCase(
                                case_id=case_id,
                                source_file=xlsx_path.name,
                                source_sheet=source_sheet,
                                row_num=row_num,
                                sequence_number=sequence_number,
                                year=year,
                                month=month,
                                date_text=date_text,
                                department=department,
                                user=user,
                                issue_text=issue_text,
                                action_text=action_text,
                                apc=apc,
                                pc_filter=pc_filter,
                                utmp=utmp,
                                sheet_title=sheet_title,
                                search_text=searchable_text,
                            )
                        )
                        case_id += 1
                        sheet_valid_cases += 1

                    report["sheet_details"].append(
                        {
                            "file": xlsx_path.name,
                            "path": relative_path,
                            "sheet": source_sheet,
                            "status": "읽음",
                            "reason": "",
                            "rows": sheet_rows,
                            "valid_cases": sheet_valid_cases,
                            "excluded_rows": sheet_excluded_rows,
                        }
                    )

                if len(sheet_names) > 1:
                    report["multi_sheet_files"] = int(report["multi_sheet_files"]) + 1
                    report["multi_sheet_details"].append(
                        {
                            "file": xlsx_path.name,
                            "path": relative_path,
                            "total_sheets": len(sheet_names),
                            "read_sheets": read_sheet_count,
                            "note": "유지보수 양식 시트만 읽음" if read_sheet_count < len(sheet_names) else "전체 시트 읽음",
                        }
                    )

                if read_sheet_count:
                    report["read_files"] = int(report["read_files"]) + 1
                else:
                    report["skipped_files"] = int(report["skipped_files"]) + 1
                    report["skipped_file_details"].append(
                        {"file": xlsx_path.name, "path": relative_path, "reason": "유지보수 양식 시트 없음"}
                    )
            finally:
                workbook.close()

        report["valid_cases"] = len(records)
        return records, report

    @staticmethod
    def _legacy_report(valid_cases: int) -> dict[str, object]:
        report = create_empty_index_report()
        report["valid_cases"] = valid_cases
        return report

    def _apply_filters(self, filters: SearchFilters) -> list[int]:
        indices: list[int] = []
        department_term = filters.department.strip().lower()
        user_term = filters.user.strip().lower()

        for idx, record in enumerate(self.records):
            if filters.year_from is not None:
                if not record.year or record.year < filters.year_from:
                    continue
            if filters.year_to is not None:
                if not record.year or record.year > filters.year_to:
                    continue
            if filters.month_from is not None:
                if not record.year or not record.month or (record.year, record.month) < filters.month_from:
                    continue
            if filters.month_to is not None:
                if not record.year or not record.month or (record.year, record.month) > filters.month_to:
                    continue
            if department_term and department_term not in record.department.lower():
                continue
            if user_term and user_term not in record.user.lower():
                continue
            if filters.require_apc and not is_checked(record.apc):
                continue
            if filters.require_pc_filter and not is_checked(record.pc_filter):
                continue
            if filters.require_utmp and not is_checked(record.utmp):
                continue
            indices.append(idx)
        return indices

    def search(
        self,
        query: str,
        filters: SearchFilters,
        top_n: int = DEFAULT_TOP_N,
    ) -> list[dict[str, object]]:
        if not self.is_ready or self.artifacts is None:
            raise RuntimeError("검색 인덱스가 아직 준비되지 않았습니다.")

        query = query.strip()
        if not query:
            return []

        candidate_indices = self._apply_filters(filters)
        if not candidate_indices:
            return []

        query_tokens = tokenize(query)
        bm25_scores = self.artifacts.bm25.score(query_tokens)
        candidate_scores = bm25_scores[candidate_indices]

        query_vector = self.artifacts.vectorizer.transform([query])
        candidate_matrix = self.artifacts.tfidf_matrix[candidate_indices]
        vector_scores = np.asarray(candidate_matrix.dot(query_vector.T).toarray()).ravel()

        bm25_norm = self._normalize_scores(candidate_scores)
        vec_norm = self._normalize_scores(vector_scores)
        final_scores = (0.45 * bm25_norm) + (0.55 * vec_norm)
        positive_indices = np.where(final_scores > 1e-8)[0]
        if positive_indices.size == 0:
            return []

        ranked = positive_indices[np.argsort(-final_scores[positive_indices])][:top_n]
        results: list[dict[str, object]] = []
        for rank, rel_idx in enumerate(ranked, start=1):
            case = self.records[candidate_indices[rel_idx]]
            results.append(
                {
                    "rank": rank,
                    "score": float(final_scores[rel_idx]),
                    "bm25": float(candidate_scores[rel_idx]),
                    "vector": float(vector_scores[rel_idx]),
                    "keyword_score": float(bm25_norm[rel_idx]),
                    "similarity_score": float(vec_norm[rel_idx]),
                    "case": case,
                }
            )
        return results

    def list_cases(self, filters: SearchFilters) -> list[dict[str, object]]:
        if not self.is_ready:
            raise RuntimeError("검색 인덱스가 아직 준비되지 않았습니다.")

        candidate_indices = self._apply_filters(filters)
        results: list[dict[str, object]] = []
        for rank, record_index in enumerate(candidate_indices, start=1):
            case = self.records[record_index]
            results.append(
                {
                    "rank": rank,
                    "score": 0.0,
                    "bm25": 0.0,
                    "vector": 0.0,
                    "keyword_score": 0.0,
                    "similarity_score": 0.0,
                    "case": case,
                    "mode": "list",
                }
            )
        return results

    @staticmethod
    def _normalize_scores(values: np.ndarray) -> np.ndarray:
        if values.size == 0:
            return values.astype(np.float32)
        max_value = float(values.max())
        min_value = float(values.min())
        if np.isclose(max_value, min_value):
            if max_value > 0:
                return np.ones_like(values, dtype=np.float32)
            return np.zeros_like(values, dtype=np.float32)
        return ((values - min_value) / (max_value - min_value)).astype(np.float32)


class MaintenanceSearchApp:
    # Tkinter 화면과 사용자 동작을 담당한다. 실제 검색/캐시는 MaintenanceSearchEngine에 위임한다.
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("1400x860")
        self.root.minsize(980, 700)

        self.engine = MaintenanceSearchEngine()
        self.queue: Queue[tuple[str, object]] = Queue()
        self.search_results: list[dict[str, object]] = []
        self.year_values: list[int] = []
        self.month_values: list[str] = []
        self._build_busy = False

        self.folder_var = tk.StringVar(value=str(Path.cwd() / "유지보수내역서 25.01~26.04"))
        self.query_var = tk.StringVar()
        self.top_n_var = tk.IntVar(value=20)
        self.year_from_var = tk.StringVar(value="전체")
        self.year_to_var = tk.StringVar(value="전체")
        self.month_from_var = tk.StringVar(value="전체")
        self.month_to_var = tk.StringVar(value="전체")
        self.department_var = tk.StringVar()
        self.user_var = tk.StringVar()
        self.apc_var = tk.BooleanVar(value=False)
        self.pc_filter_var = tk.BooleanVar(value=False)
        self.utmp_var = tk.BooleanVar(value=False)
        self.status_var = tk.StringVar(value="인덱스를 불러오거나 새로 구축하세요.")
        self.index_notice_var = tk.StringVar(value="")
        self._last_stale_warning_key = ""

        self._build_ui()
        self._show_query_hint()
        self._try_load_existing_index()
        self._schedule_manifest_check()

    def _build_ui(self) -> None:
        style = ttk.Style()
        try:
            theme_names = style.theme_names()
            if sys.platform == "win32" and "vista" in theme_names:
                style.theme_use("vista")
            elif sys.platform == "darwin" and "aqua" in theme_names:
                style.theme_use("aqua")
            elif "clam" in theme_names:
                style.theme_use("clam")
        except Exception:
            pass
        default_font = ("맑은 고딕", 10)
        style.configure(".", font=default_font)
        style.configure("Search.TEntry", foreground="#111111")
        style.configure("SearchHint.TEntry", foreground="#777777")
        style.configure("TLabelframe.Label", font=("맑은 고딕", 10, "bold"))
        style.configure("Warning.TLabel", foreground="#B91C1C")
        style.configure("Treeview", rowheight=24)
        style.configure("Treeview.Heading", font=("맑은 고딕", 10, "bold"))

        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(1, weight=0)

        top = ttk.Frame(self.root, padding=10)
        top.grid(row=0, column=0, sticky="ew")
        top.columnconfigure(1, weight=1)

        ttk.Label(top, text="데이터 폴더").grid(row=0, column=0, sticky="w")
        folder_entry = ttk.Entry(top, textvariable=self.folder_var)
        folder_entry.grid(row=0, column=1, sticky="ew", padx=(8, 8))
        ttk.Button(top, text="찾기", command=self._browse_folder).grid(row=0, column=2, padx=(0, 6))
        self.build_button = ttk.Button(top, text="인덱스 구축", command=self._prepare_index)
        self.build_button.grid(row=0, column=3, padx=(0, 6))
        ttk.Button(top, text="인덱스 리포트 보기", command=self._show_index_report).grid(row=0, column=4, padx=(0, 6))
        ttk.Button(top, text="새 내역 추가", command=self._show_add_case_dialog).grid(row=0, column=5)
        ttk.Label(top, textvariable=self.index_notice_var, style="Warning.TLabel").grid(
            row=1, column=1, columnspan=5, sticky="w", padx=(8, 0), pady=(6, 0)
        )

        search = ttk.LabelFrame(self.root, text="검색", padding=10)
        search.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 10))
        search.columnconfigure(0, weight=1)

        query_row = ttk.Frame(search)
        query_row.grid(row=0, column=0, sticky="ew")
        query_row.columnconfigure(1, weight=1)

        ttk.Label(query_row, text="검색어").grid(row=0, column=0, sticky="w")
        self.query_entry = ttk.Entry(query_row, textvariable=self.query_var, style="Search.TEntry")
        self.query_entry.grid(row=0, column=1, sticky="ew", padx=(8, 0))
        self.query_entry.bind("<FocusIn>", self._hide_query_hint)
        self.query_entry.bind("<FocusOut>", self._show_query_hint)
        self.query_entry.bind("<Return>", lambda _: self._run_search())

        action_row = ttk.Frame(search)
        action_row.grid(row=1, column=0, sticky="ew", pady=(8, 0))
        ttk.Label(action_row, text="결과 수").pack(side=tk.LEFT)
        ttk.Spinbox(action_row, from_=5, to=100, textvariable=self.top_n_var, width=6).pack(side=tk.LEFT, padx=(8, 12))
        actions = ttk.Frame(action_row)
        actions.pack(side=tk.LEFT)
        ttk.Button(actions, text="검색", command=self._run_search).pack(side=tk.LEFT)
        ttk.Button(actions, text="전체 보기", command=self._show_all_cases).pack(side=tk.LEFT, padx=(6, 0))

        filter_frame = ttk.Frame(search)
        filter_frame.grid(row=2, column=0, sticky="ew", pady=(10, 8))
        for col in range(8):
            filter_frame.columnconfigure(col, weight=0)

        ttk.Label(filter_frame, text="시작 연도").grid(row=0, column=0, sticky="w")
        self.year_from_combo = ttk.Combobox(filter_frame, textvariable=self.year_from_var, values=["전체"], width=9, state="readonly")
        self.year_from_combo.grid(row=0, column=1, sticky="w", padx=(6, 12))
        self.year_from_combo.bind("<<ComboboxSelected>>", lambda _: self._normalize_filter_range("year_from"))

        ttk.Label(filter_frame, text="종료 연도").grid(row=0, column=2, sticky="w")
        self.year_to_combo = ttk.Combobox(filter_frame, textvariable=self.year_to_var, values=["전체"], width=9, state="readonly")
        self.year_to_combo.grid(row=0, column=3, sticky="w", padx=(6, 12))
        self.year_to_combo.bind("<<ComboboxSelected>>", lambda _: self._normalize_filter_range("year_to"))

        ttk.Label(filter_frame, text="시작 월").grid(row=0, column=4, sticky="w")
        self.month_from_combo = ttk.Combobox(filter_frame, textvariable=self.month_from_var, values=["전체"], width=9, state="readonly")
        self.month_from_combo.grid(row=0, column=5, sticky="w", padx=(6, 12))
        self.month_from_combo.bind("<<ComboboxSelected>>", lambda _: self._normalize_filter_range("month_from"))

        ttk.Label(filter_frame, text="종료 월").grid(row=0, column=6, sticky="w")
        self.month_to_combo = ttk.Combobox(filter_frame, textvariable=self.month_to_var, values=["전체"], width=9, state="readonly")
        self.month_to_combo.grid(row=0, column=7, sticky="w", padx=(6, 0))
        self.month_to_combo.bind("<<ComboboxSelected>>", lambda _: self._normalize_filter_range("month_to"))

        ttk.Label(filter_frame, text="부서 포함").grid(row=1, column=0, sticky="w", pady=(8, 0))
        ttk.Entry(filter_frame, textvariable=self.department_var, width=16).grid(
            row=1, column=1, sticky="w", padx=(6, 16), pady=(8, 0)
        )

        ttk.Label(filter_frame, text="사용자 포함").grid(row=1, column=2, sticky="w", pady=(8, 0))
        ttk.Entry(filter_frame, textvariable=self.user_var, width=14).grid(
            row=1, column=3, sticky="w", padx=(6, 16), pady=(8, 0)
        )

        flags = ttk.Frame(filter_frame)
        flags.grid(row=1, column=4, columnspan=4, sticky="w", pady=(8, 0))
        ttk.Checkbutton(flags, text="APC", variable=self.apc_var).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Checkbutton(flags, text="PC filter", variable=self.pc_filter_var).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Checkbutton(flags, text="UTMP", variable=self.utmp_var).pack(side=tk.LEFT)

        body = ttk.Panedwindow(self.root, orient=tk.VERTICAL)
        body.grid(row=2, column=0, sticky="nsew", padx=10, pady=(0, 10))
        self.root.rowconfigure(2, weight=1)

        results_frame = ttk.Labelframe(body, text="유사 사례 목록", padding=6)
        detail_frame = ttk.Labelframe(body, text="선택 항목 상세", padding=6)
        body.add(results_frame, weight=3)
        body.add(detail_frame, weight=2)

        results_frame.rowconfigure(0, weight=1)
        results_frame.columnconfigure(0, weight=1)
        columns = ("rank", "score", "year", "date", "dept", "user", "issue", "action", "file")
        self.tree = ttk.Treeview(results_frame, columns=columns, show="headings", height=16)
        headings = {
            "rank": "순위",
            "score": "관련도",
            "year": "연도",
            "date": "날짜",
            "dept": "부서",
            "user": "사용자",
            "issue": "장애내용",
            "action": "조치내용",
            "file": "원본파일",
        }
        widths = {
            "rank": 45,
            "score": 70,
            "year": 55,
            "date": 85,
            "dept": 130,
            "user": 90,
            "issue": 320,
            "action": 430,
            "file": 190,
        }
        for col in columns:
            self.tree.heading(col, text=headings[col])
            self.tree.column(col, width=widths[col], anchor="w", stretch=col in {"issue", "action", "file"})
        vsb = ttk.Scrollbar(results_frame, orient=tk.VERTICAL, command=self.tree.yview)
        hsb = ttk.Scrollbar(results_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscroll=vsb.set, xscroll=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        self.tree.bind("<<TreeviewSelect>>", self._show_selected_detail)

        ttk.Button(results_frame, text="검색 결과 엑셀 저장", command=self._export_results).grid(
            row=2, column=0, sticky="e", pady=(8, 0)
        )

        detail_frame.rowconfigure(1, weight=1)
        detail_frame.columnconfigure(0, weight=1)

        detail_meta = ttk.Frame(detail_frame)
        detail_meta.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        for col in range(4):
            detail_meta.columnconfigure(col, weight=1)
        self.detail_rank_var = tk.StringVar(value="사례를 선택하세요.")
        self.detail_score_var = tk.StringVar(value="")
        self.detail_date_var = tk.StringVar(value="")
        self.detail_person_var = tk.StringVar(value="")
        self.detail_flags_var = tk.StringVar(value="")
        ttk.Label(detail_meta, textvariable=self.detail_rank_var, font=("맑은 고딕", 10, "bold")).grid(
            row=0, column=0, sticky="w", padx=(0, 12)
        )
        ttk.Label(detail_meta, textvariable=self.detail_score_var).grid(row=0, column=1, sticky="w", padx=(0, 12))
        ttk.Label(detail_meta, textvariable=self.detail_date_var).grid(row=0, column=2, sticky="w", padx=(0, 12))
        ttk.Label(detail_meta, textvariable=self.detail_person_var).grid(row=0, column=3, sticky="w")
        ttk.Label(detail_meta, textvariable=self.detail_flags_var).grid(row=1, column=0, columnspan=4, sticky="w", pady=(4, 0))

        detail_content = ttk.Frame(detail_frame)
        detail_content.grid(row=1, column=0, sticky="nsew")
        detail_content.rowconfigure(1, weight=1)
        detail_content.columnconfigure(0, weight=1, uniform="detail")
        detail_content.columnconfigure(1, weight=1, uniform="detail")

        ttk.Label(detail_content, text="장애내용", font=("맑은 고딕", 10, "bold")).grid(
            row=0, column=0, sticky="w", pady=(0, 4)
        )
        ttk.Label(detail_content, text="조치내용", font=("맑은 고딕", 10, "bold")).grid(
            row=0, column=1, sticky="w", padx=(8, 0), pady=(0, 4)
        )

        issue_frame = ttk.Frame(detail_content)
        issue_frame.grid(row=1, column=0, sticky="nsew", padx=(0, 8))
        issue_frame.rowconfigure(0, weight=1)
        issue_frame.columnconfigure(0, weight=1)
        self.detail_issue_text = tk.Text(
            issue_frame,
            wrap="word",
            height=7,
            bg="#FFFFFF",
            fg="#1F2937",
            insertbackground="#1F2937",
            relief=tk.FLAT,
            padx=10,
            pady=8,
            font=("맑은 고딕", 10),
        )
        self.detail_issue_text.grid(row=0, column=0, sticky="nsew")
        issue_scroll = ttk.Scrollbar(issue_frame, orient=tk.VERTICAL, command=self.detail_issue_text.yview)
        self.detail_issue_text.configure(yscrollcommand=issue_scroll.set)
        issue_scroll.grid(row=0, column=1, sticky="ns")

        action_frame = ttk.Frame(detail_content)
        action_frame.grid(row=1, column=1, sticky="nsew")
        action_frame.rowconfigure(0, weight=1)
        action_frame.columnconfigure(0, weight=1)
        self.detail_action_text = tk.Text(
            action_frame,
            wrap="word",
            height=7,
            bg="#FFFFFF",
            fg="#1F2937",
            insertbackground="#1F2937",
            relief=tk.FLAT,
            padx=10,
            pady=8,
            font=("맑은 고딕", 10),
        )
        self.detail_action_text.grid(row=0, column=0, sticky="nsew")
        action_scroll = ttk.Scrollbar(action_frame, orient=tk.VERTICAL, command=self.detail_action_text.yview)
        self.detail_action_text.configure(yscrollcommand=action_scroll.set)
        action_scroll.grid(row=0, column=1, sticky="ns")

        self.detail_source_var = tk.StringVar(value="")
        self.detail_source_label = ttk.Label(detail_frame, textvariable=self.detail_source_var)
        self.detail_source_label.grid(row=2, column=0, sticky="ew", pady=(8, 0))
        detail_frame.bind(
            "<Configure>",
            lambda event: self.detail_source_label.configure(wraplength=max(300, event.width - 24)),
            add="+",
        )
        detail_actions = ttk.Frame(detail_frame)
        detail_actions.grid(row=3, column=0, sticky="e", pady=(8, 0))
        ttk.Button(detail_actions, text="원본 열기", command=self._open_selected_case_source).pack(side=tk.LEFT)
        ttk.Button(detail_actions, text="선택 항목 수정", command=self._show_edit_case_dialog).pack(side=tk.LEFT, padx=(6, 0))
        self._set_detail_empty_message("검색 결과에서 사례를 선택하면 상세 내용이 표시됩니다.")

        bottom = ttk.Frame(self.root, padding=(10, 0, 10, 10))
        bottom.grid(row=3, column=0, sticky="ew")
        bottom.columnconfigure(0, weight=1)
        self.progress = ttk.Progressbar(bottom, mode="indeterminate")
        self.progress.grid(row=0, column=0, sticky="ew", padx=(0, 10))
        ttk.Label(bottom, textvariable=self.status_var).grid(row=1, column=0, sticky="w", pady=(6, 0))

    def _browse_folder(self) -> None:
        selected = filedialog.askdirectory(title="유지보수 엑셀 폴더 선택", initialdir=self.folder_var.get() or str(Path.cwd()))
        if selected:
            self.folder_var.set(selected)
            self.status_var.set(f"폴더 선택 완료: {selected}")
            self._load_index()

    def _show_query_hint(self, event: object | None = None) -> None:
        if not self.query_var.get().strip():
            self.query_var.set(QUERY_HINT)
            self.query_entry.configure(style="SearchHint.TEntry")

    def _hide_query_hint(self, event: object | None = None) -> None:
        if self.query_var.get() == QUERY_HINT:
            self.query_var.set("")
            self.query_entry.configure(style="Search.TEntry")

    def _set_index_notice(self, stale: bool, detail: str = "") -> None:
        if stale:
            suffix = f" ({detail})" if detail else ""
            self.index_notice_var.set(f"데이터 변경 감지: [인덱스 구축]을 눌러 새로 구축하세요.{suffix}")
            self.build_button.configure(text="인덱스 새로 구축")
        else:
            self.index_notice_var.set("")
            self.build_button.configure(text="인덱스 구축")
            self._last_stale_warning_key = ""

    def _warn_stale_index_once(self, folder: Path) -> None:
        warning_key = f"{folder.resolve()}|{self.engine.index_stale_message}"
        if self._last_stale_warning_key == warning_key:
            return
        self._last_stale_warning_key = warning_key
        detail = f"\n\n변경 내용: {self.engine.index_stale_message}" if self.engine.index_stale_message else ""
        messagebox.showwarning(
            APP_TITLE,
            "데이터 폴더의 Excel 파일이 변경되었습니다.\n"
            "[인덱스 구축]을 눌러 새로 구축해야 최신 검색 결과가 반영됩니다."
            f"{detail}",
        )

    def _schedule_manifest_check(self) -> None:
        self.root.after(MANIFEST_CHECK_INTERVAL_MS, self._check_source_files_changed)

    def _check_source_files_changed(self) -> None:
        try:
            if not self._build_busy and self.engine.is_ready and self.engine.loaded_folder and self.engine.index_report:
                folder = self.engine.loaded_folder
                if folder.exists():
                    stale, message = compare_source_file_manifest(folder, self.engine.index_report)
                    self.engine.index_stale = stale
                    self.engine.index_stale_message = message
                    if stale:
                        self._set_index_notice(True, message)
                        self.status_var.set("데이터 변경 감지: 인덱스를 새로 구축하세요.")
                        self._warn_stale_index_once(folder)
                    else:
                        self._set_index_notice(False)
        except Exception:
            pass
        finally:
            self._schedule_manifest_check()

    def _prepare_index(self) -> None:
        folder = Path(self.folder_var.get()).expanduser()
        if not folder.exists():
            messagebox.showerror(APP_TITLE, "데이터 폴더를 찾을 수 없습니다.")
            return
        if self._build_busy:
            return
        self._start_build(folder)

    def _start_build(self, folder: Optional[Path] = None) -> None:
        folder = folder or Path(self.folder_var.get()).expanduser()
        if not folder.exists():
            messagebox.showerror(APP_TITLE, "데이터 폴더를 찾을 수 없습니다.")
            return
        self._set_busy(True)
        self._build_busy = True
        self.status_var.set("인덱스를 구축 중입니다. 잠시만 기다려 주세요.")
        threading.Thread(target=self._build_worker, args=(folder,), daemon=True).start()
        self.root.after(100, self._poll_queue)

    def _build_worker(self, folder: Path) -> None:
        try:
            count, years, report = self.engine.build_from_folder(folder)
            self.queue.put(("build_done", (count, years, folder, report)))
        except Exception as exc:
            self.queue.put(("error", (exc, traceback.format_exc())))

    def _load_index(self) -> None:
        folder = Path(self.folder_var.get()).expanduser()
        if not folder.exists():
            self.status_var.set("데이터 폴더가 없습니다.")
            self._set_index_notice(False)
            return
        try:
            if self.engine.load_folder(folder):
                self.year_values = self.engine.artifacts.years if self.engine.artifacts else []
                self.month_values = self._get_month_values()
                self._refresh_filter_combos()
                if self.engine.index_stale:
                    self._set_index_notice(True, self.engine.index_stale_message)
                    self.status_var.set("데이터 변경 감지: 인덱스를 새로 구축하세요.")
                    self._warn_stale_index_once(folder)
                else:
                    self._set_index_notice(False)
                    self.status_var.set(f"인덱스 로드 완료: {len(self.engine.records)}건")
            else:
                self._set_index_notice(False)
                self.status_var.set("저장된 인덱스가 없습니다. 인덱스 구축을 실행하세요.")
        except Exception as exc:
            messagebox.showerror(APP_TITLE, f"인덱스 로드 실패\n\n{exc}")

    def _try_load_existing_index(self) -> None:
        self._load_index()

    def _poll_queue(self) -> None:
        try:
            kind, payload = self.queue.get_nowait()
        except Empty:
            if self._build_busy:
                self.root.after(100, self._poll_queue)
            return

        if kind == "build_done":
            count, years, folder, report = payload
            self.year_values = years
            self.month_values = self._get_month_values()
            self._refresh_filter_combos()
            self._last_stale_warning_key = ""
            self._set_index_notice(False)
            self.status_var.set(f"인덱스 구축 완료: {count}건 / 폴더: {folder}")
            messagebox.showinfo(APP_TITLE, self._format_build_summary(report))
        elif kind == "error":
            exc, tb = payload
            messagebox.showerror(APP_TITLE, f"작업 중 오류가 발생했습니다.\n\n{exc}\n\n{tb}")
            self.status_var.set("오류가 발생했습니다.")
        self._build_busy = False
        self._set_busy(False)

    def _format_build_summary(self, report: dict[str, object]) -> str:
        return (
            "인덱스 구축 완료\n"
            f"읽은 파일: {report.get('read_files', 0)}개 / 전체 {report.get('total_files', 0)}개\n"
            f"유효 사례: {report.get('valid_cases', 0)}건\n"
            f"제외 행: {report.get('excluded_rows', 0)}건\n"
            f"제외 파일: {report.get('excluded_files', 0)}개\n"
            f"다중 시트 파일: {report.get('multi_sheet_files', 0)}개\n"
            "자세한 내용은 [인덱스 리포트 보기]에서 확인하세요."
        )

    def _show_index_report(self) -> None:
        report = self.engine.index_report
        if report is None:
            folder = Path(self.folder_var.get()).expanduser()
            if folder.exists():
                report = MaintenanceRepository(folder).load_report()
                self.engine.index_report = report
        if report is None:
            messagebox.showwarning(APP_TITLE, "확인할 인덱스 리포트가 없습니다.")
            return

        window = tk.Toplevel(self.root)
        window.title("인덱스 리포트")
        window.geometry("780x520")
        window.minsize(680, 440)

        notebook = ttk.Notebook(window)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        summary_rows = [
            ("전체 파일", report.get("total_files", 0)),
            ("읽은 파일", report.get("read_files", 0)),
            ("제외 파일", report.get("excluded_files", 0)),
            ("건너뛴 파일", report.get("skipped_files", 0)),
            ("총 행 수", report.get("total_rows", 0)),
            ("유효 사례", report.get("valid_cases", 0)),
            ("제외 행", report.get("excluded_rows", 0)),
            ("다중 시트 파일", report.get("multi_sheet_files", 0)),
        ]
        summary_frame = ttk.Frame(notebook, padding=8)
        notebook.add(summary_frame, text="요약")
        self._populate_tree(summary_frame, ("item", "value"), {"item": "항목", "value": "값"}, summary_rows)

        excluded_rows = [
            (item.get("file", ""), "제외", item.get("reason", ""))
            for item in report.get("excluded_file_details", [])
        ]
        excluded_rows.extend(
            (item.get("file", ""), "건너뜀", item.get("reason", ""))
            for item in report.get("skipped_file_details", [])
        )
        excluded_frame = ttk.Frame(notebook, padding=8)
        notebook.add(excluded_frame, text="제외 파일")
        self._populate_tree(
            excluded_frame,
            ("file", "status", "reason"),
            {"file": "파일명", "status": "구분", "reason": "사유"},
            excluded_rows,
        )

        multi_sheet_rows = [
            (
                item.get("file", ""),
                item.get("total_sheets", 0),
                item.get("read_sheets", 0),
                item.get("note", ""),
            )
            for item in report.get("multi_sheet_details", [])
        ]
        multi_frame = ttk.Frame(notebook, padding=8)
        notebook.add(multi_frame, text="다중 시트")
        self._populate_tree(
            multi_frame,
            ("file", "total", "read", "note"),
            {"file": "파일명", "total": "전체 시트 수", "read": "읽은 시트 수", "note": "비고"},
            multi_sheet_rows,
        )

        quality_rows = [
            ("유효 사례", report.get("valid_cases", 0)),
            ("제외 행", report.get("excluded_rows", 0)),
            ("부서 누락", report.get("missing_departments", 0)),
            ("사용자 누락", report.get("missing_users", 0)),
            ("APC 누락", report.get("missing_apc", 0)),
            ("PC filter 누락", report.get("missing_pc_filter", 0)),
            ("UTMP 누락", report.get("missing_utmp", 0)),
        ]
        quality_frame = ttk.Frame(notebook, padding=8)
        notebook.add(quality_frame, text="데이터 품질")
        quality_frame.columnconfigure(0, weight=1)
        quality_summary = ttk.Frame(quality_frame)
        quality_summary.grid(row=0, column=0, sticky="ew")
        self._populate_tree(quality_summary, ("item", "value"), {"item": "항목", "value": "건수"}, quality_rows)

        detail_visible_var = tk.BooleanVar(value=False)
        detail_rows = self._get_quality_detail_rows(report)
        detail_controls = ttk.Frame(quality_frame)
        detail_controls.grid(row=1, column=0, sticky="ew", pady=(8, 4))
        detail_frame = ttk.Frame(quality_frame)
        detail_frame.grid(row=2, column=0, sticky="nsew")
        quality_frame.rowconfigure(2, weight=1)
        detail_tree, detail_by_iid = self._populate_report_detail_tree(detail_frame, detail_rows)
        detail_frame.grid_remove()
        detail_status_var = tk.StringVar(value="상세 행을 선택하면 원본 파일 위치를 열 수 있습니다.")

        def toggle_detail() -> None:
            if detail_visible_var.get():
                detail_frame.grid()
            else:
                detail_frame.grid_remove()

        def update_detail_action(_: object = None) -> None:
            selection = detail_tree.selection()
            can_open = bool(selection and selection[0] in detail_by_iid)
            if can_open:
                open_button.state(["!disabled"])
                detail_status_var.set("선택한 행의 원본 파일 위치를 열 수 있습니다.")
            else:
                open_button.state(["disabled"])
                detail_status_var.set("상세 행을 선택하면 원본 파일 위치를 열 수 있습니다.")

        ttk.Checkbutton(
            detail_controls,
            text="상세 행 위치 표시",
            variable=detail_visible_var,
            command=toggle_detail,
        ).pack(side=tk.LEFT)
        open_button = ttk.Button(
            detail_controls,
            text="선택 위치 열기",
            command=lambda: self._open_selected_report_detail(detail_tree, detail_by_iid, detail_status_var),
        )
        open_button.pack(side=tk.LEFT, padx=(8, 0))
        open_button.state(["disabled"])
        ttk.Label(detail_controls, textvariable=detail_status_var).pack(side=tk.LEFT, padx=(10, 0))
        detail_tree.bind("<<TreeviewSelect>>", update_detail_action, add="+")

    def _populate_tree(
        self,
        parent: ttk.Frame,
        columns: tuple[str, ...],
        headings: dict[str, str],
        rows: list[tuple[object, ...]],
    ) -> None:
        parent.rowconfigure(0, weight=1)
        parent.columnconfigure(0, weight=1)
        tree = ttk.Treeview(parent, columns=columns, show="headings")
        for column in columns:
            tree.heading(column, text=headings[column])
            tree.column(column, width=180 if column == "file" else 120, anchor="w")
        if rows:
            for row in rows:
                tree.insert("", tk.END, values=row)
        else:
            tree.insert("", tk.END, values=("표시할 내용 없음",) + ("",) * (len(columns) - 1))

        vsb = ttk.Scrollbar(parent, orient=tk.VERTICAL, command=tree.yview)
        hsb = ttk.Scrollbar(parent, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscroll=vsb.set, xscroll=hsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

    def _get_quality_detail_rows(self, report: dict[str, object]) -> list[dict[str, object]]:
        details: list[dict[str, object]] = []
        details.extend(report.get("quality_details", []))
        details.extend(report.get("excluded_row_details", []))
        return details

    def _populate_report_detail_tree(
        self,
        parent: ttk.Frame,
        rows: list[dict[str, object]],
    ) -> tuple[ttk.Treeview, dict[str, dict[str, object]]]:
        parent.rowconfigure(0, weight=1)
        parent.columnconfigure(0, weight=1)
        columns = ("type", "file", "sheet", "row", "reason", "issue", "action")
        headings = {
            "type": "구분",
            "file": "파일명",
            "sheet": "시트",
            "row": "행",
            "reason": "사유",
            "issue": "장애내용",
            "action": "조치내용",
        }
        widths = {
            "type": 120,
            "file": 170,
            "sheet": 110,
            "row": 60,
            "reason": 180,
            "issue": 220,
            "action": 260,
        }
        tree = ttk.Treeview(parent, columns=columns, show="headings", height=10)
        for column in columns:
            tree.heading(column, text=headings[column])
            tree.column(column, width=widths[column], anchor="w")

        detail_by_iid: dict[str, dict[str, object]] = {}
        if rows:
            for detail in rows:
                iid = tree.insert(
                    "",
                    tk.END,
                    values=(
                        detail.get("type", ""),
                        detail.get("file", ""),
                        detail.get("sheet", ""),
                        detail.get("row", ""),
                        detail.get("reason", ""),
                        self._shorten(str(detail.get("issue", "")), 30),
                        self._shorten(str(detail.get("action", "")), 36),
                    ),
                )
                detail_by_iid[iid] = detail
        else:
            tree.insert("", tk.END, values=("표시할 상세 행 없음", "", "", "", "", "", ""))

        tree.bind("<Double-1>", lambda _: self._open_selected_report_detail(tree, detail_by_iid))
        vsb = ttk.Scrollbar(parent, orient=tk.VERTICAL, command=tree.yview)
        hsb = ttk.Scrollbar(parent, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscroll=vsb.set, xscroll=hsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        return tree, detail_by_iid

    def _open_selected_report_detail(
        self,
        tree: ttk.Treeview,
        detail_by_iid: dict[str, dict[str, object]],
        status_var: Optional[tk.StringVar] = None,
    ) -> None:
        selection = tree.selection()
        if not selection:
            if status_var is not None:
                status_var.set("상세 행을 먼저 선택하세요.")
            else:
                messagebox.showwarning(APP_TITLE, "먼저 상세 행을 선택하세요.")
            return
        detail = detail_by_iid.get(selection[0])
        if not detail:
            if status_var is not None:
                status_var.set("열 수 있는 상세 행이 아닙니다.")
            return
        self._open_report_location(detail)

    def _open_report_location(self, detail: dict[str, object]) -> None:
        path_text = str(detail.get("path") or detail.get("file") or "")
        if not path_text:
            messagebox.showwarning(APP_TITLE, "원본 파일 경로가 없습니다.")
            return
        path = Path(path_text)
        if not path.is_absolute():
            base = self.engine.loaded_folder or Path(self.folder_var.get()).expanduser()
            path = base / path
        if not path.exists():
            messagebox.showwarning(APP_TITLE, f"원본 파일을 찾을 수 없습니다.\n\n{path}")
            return

        sheet_name = str(detail.get("sheet") or "")
        row_num = detail.get("row")
        self._open_excel_file_location(path, sheet_name, row_num)

    # 원본 위치 열기는 Windows Excel COM이 가능하면 행까지 이동하고, 실패하면 파일 열기로 대체한다.
    def _open_excel_file_location(self, path: Path, sheet_name: str = "", row_num: object = None) -> None:
        if sys.platform == "win32" and sheet_name and row_num:
            try:
                import win32com.client  # type: ignore[import-not-found]

                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = True
                workbook = excel.Workbooks.Open(str(path.resolve()))
                worksheet = workbook.Worksheets(sheet_name)
                worksheet.Activate()
                worksheet.Rows(int(row_num)).Select()
                return
            except Exception:
                pass

        try:
            if sys.platform == "win32":
                os.startfile(path)  # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                subprocess.Popen(["open", str(path)])
            else:
                subprocess.Popen(["xdg-open", str(path)])
        except Exception as exc:
            messagebox.showerror(APP_TITLE, f"원본 파일 열기 실패\n\n{exc}")

    # 검색 결과의 원본 Excel 행을 직접 열거나 수정하는 기능 묶음.
    def _get_selected_case(self) -> Optional[MaintenanceCase]:
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning(APP_TITLE, "먼저 검색 결과에서 항목을 선택하세요.")
            return None
        idx = self.tree.index(selection[0])
        if idx >= len(self.search_results):
            messagebox.showwarning(APP_TITLE, "선택한 항목을 찾을 수 없습니다.")
            return None
        return self.search_results[idx]["case"]

    def _resolve_case_path(self, case: MaintenanceCase) -> Optional[Path]:
        base = self.engine.loaded_folder or Path(self.folder_var.get()).expanduser()
        source = Path(case.source_file)
        candidates: list[Path] = []
        if source.is_absolute():
            candidates.append(source)
        else:
            candidates.append(base / source)
            if base.exists():
                candidates.extend(sorted(base.rglob(source.name)))

        for candidate in candidates:
            if candidate.exists() and candidate.is_file():
                year, month = parse_year_month_from_path(candidate)
                if case.year and case.month and (year, month) == (case.year, case.month):
                    return candidate
        for candidate in candidates:
            if candidate.exists() and candidate.is_file():
                return candidate
        return None

    def _open_selected_case_source(self) -> None:
        case = self._get_selected_case()
        if case is None:
            return
        path = self._resolve_case_path(case)
        if path is None:
            messagebox.showwarning(APP_TITLE, f"원본 파일을 찾을 수 없습니다.\n\n{case.source_file}")
            return
        self._open_excel_file_location(path, case.source_sheet, case.row_num)

    def _show_add_case_dialog(self) -> None:
        folder = Path(self.folder_var.get()).expanduser()
        if not folder.exists():
            messagebox.showwarning(APP_TITLE, "먼저 데이터 폴더를 선택하세요.")
            return
        month_values = self._get_available_source_months()
        if not month_values:
            messagebox.showwarning(APP_TITLE, "유지보수 엑셀 파일을 찾을 수 없습니다.")
            return
        initial_month = month_values[-1]
        self._show_case_editor(
            title="새 내역 추가",
            initial={
                "sequence": self._get_next_sequence_for_month(initial_month),
                "date": "",
                "department": "",
                "user": "",
                "issue": "",
                "action": "",
                "apc": True,
                "pc_filter": True,
                "utmp": True,
            },
            month_values=month_values,
            initial_month=initial_month,
            on_save=lambda values, month: self._save_new_case(month, values),
        )

    def _show_edit_case_dialog(self) -> None:
        case = self._get_selected_case()
        if case is None:
            return
        self._show_case_editor(
            title="선택 항목 수정",
            initial={
                "sequence": case.sequence_number,
                "date": case.date_text,
                "department": case.department,
                "user": case.user,
                "issue": case.issue_text,
                "action": case.action_text,
                "apc": is_checked(case.apc),
                "pc_filter": is_checked(case.pc_filter),
                "utmp": is_checked(case.utmp),
            },
            source_text=f"{case.source_file} / {case.source_sheet} / {case.row_num}행",
            on_save=lambda values, _month: self._save_existing_case(case, values),
        )

    def _show_case_editor(
        self,
        title: str,
        initial: dict[str, object],
        on_save: Callable[[dict[str, str], str], tuple[Path, str]],
        month_values: Optional[list[str]] = None,
        initial_month: str = "",
        source_text: str = "",
    ) -> None:
        window = tk.Toplevel(self.root)
        window.title(title)
        window.geometry("620x560")
        window.minsize(560, 500)
        window.transient(self.root)
        window.grab_set()
        window.columnconfigure(1, weight=1)
        row = 0

        month_var = tk.StringVar(value=initial_month)
        sequence_var = tk.StringVar(value=str(initial.get("sequence", "")))
        if month_values is not None:
            months_by_year: dict[str, list[str]] = {}
            for value in month_values:
                try:
                    year_text, month_text = value.split("-", 1)
                except ValueError:
                    continue
                months_by_year.setdefault(year_text, []).append(month_text)
            for months in months_by_year.values():
                months.sort()

            initial_year, initial_month_only = initial_month.split("-", 1)
            year_var = tk.StringVar(value=initial_year)
            month_only_var = tk.StringVar(value=initial_month_only)
            target_file_var = tk.StringVar()

            def update_target_file() -> None:
                if not year_var.get() or not month_only_var.get():
                    month_var.set("")
                    target_file_var.set("저장 파일: -")
                    return
                selected_month = f"{year_var.get()}-{month_only_var.get()}"
                month_var.set(selected_month)
                sequence_var.set(self._get_next_sequence_for_month(selected_month))
                path = self._find_workbook_for_month(selected_month)
                if path is None:
                    target_file_var.set("저장 파일: 해당 연월의 유지보수 엑셀 파일을 찾을 수 없습니다.")
                else:
                    base = Path(self.folder_var.get()).expanduser()
                    try:
                        display_path = path.relative_to(base)
                    except ValueError:
                        display_path = path
                    target_file_var.set(f"저장 파일: {display_path}")

            def update_month_choices() -> None:
                values = months_by_year.get(year_var.get(), [])
                month_combo.configure(values=values)
                if month_only_var.get() not in values:
                    month_only_var.set(values[-1] if values else "")
                update_target_file()

            ttk.Label(window, text="대상 연월").grid(row=row, column=0, sticky="w", padx=12, pady=(12, 6))
            target_frame = ttk.Frame(window)
            target_frame.grid(row=row, column=1, sticky="w", padx=12, pady=(12, 6))
            year_combo = ttk.Combobox(
                target_frame,
                textvariable=year_var,
                values=sorted(months_by_year),
                state="readonly",
                width=8,
            )
            year_combo.pack(side=tk.LEFT)
            ttk.Label(target_frame, text="년").pack(side=tk.LEFT, padx=(4, 10))
            month_combo = ttk.Combobox(target_frame, textvariable=month_only_var, state="readonly", width=6)
            month_combo.pack(side=tk.LEFT)
            ttk.Label(target_frame, text="월").pack(side=tk.LEFT, padx=(4, 0))
            year_combo.bind("<<ComboboxSelected>>", lambda _: update_month_choices())
            month_combo.bind("<<ComboboxSelected>>", lambda _: update_target_file())
            row += 1

            ttk.Label(window, text="저장 위치").grid(row=row, column=0, sticky="w", padx=12, pady=6)
            target_file_label = ttk.Label(window, textvariable=target_file_var)
            target_file_label.grid(row=row, column=1, sticky="ew", padx=12, pady=6)
            target_file_label.bind(
                "<Configure>",
                lambda event: target_file_label.configure(wraplength=max(320, event.width - 8)),
                add="+",
            )
            update_month_choices()
            row += 1
        elif source_text:
            ttk.Label(window, text="원본").grid(row=row, column=0, sticky="w", padx=12, pady=(12, 6))
            ttk.Label(window, text=source_text).grid(row=row, column=1, sticky="w", padx=12, pady=(12, 6))
            row += 1

        date_var = tk.StringVar(value=str(initial.get("date", "")))
        department_var = tk.StringVar(value=str(initial.get("department", "")))
        user_var = tk.StringVar(value=str(initial.get("user", "")))
        apc_var = tk.BooleanVar(value=bool(initial.get("apc", False)))
        pc_filter_var = tk.BooleanVar(value=bool(initial.get("pc_filter", False)))
        utmp_var = tk.BooleanVar(value=bool(initial.get("utmp", False)))

        ttk.Label(window, text="순번").grid(row=row, column=0, sticky="w", padx=12, pady=6)
        ttk.Entry(window, textvariable=sequence_var, width=12).grid(row=row, column=1, sticky="w", padx=12, pady=6)
        row += 1
        ttk.Label(window, text="날짜").grid(row=row, column=0, sticky="w", padx=12, pady=6)
        ttk.Entry(window, textvariable=date_var).grid(row=row, column=1, sticky="ew", padx=12, pady=6)
        row += 1
        ttk.Label(window, text="부서").grid(row=row, column=0, sticky="w", padx=12, pady=6)
        ttk.Entry(window, textvariable=department_var).grid(row=row, column=1, sticky="ew", padx=12, pady=6)
        row += 1
        ttk.Label(window, text="사용자").grid(row=row, column=0, sticky="w", padx=12, pady=6)
        ttk.Entry(window, textvariable=user_var).grid(row=row, column=1, sticky="ew", padx=12, pady=6)
        row += 1

        ttk.Label(window, text="장애내용").grid(row=row, column=0, sticky="nw", padx=12, pady=6)
        issue_text = tk.Text(window, height=5, wrap="word", font=("맑은 고딕", 10))
        issue_text.grid(row=row, column=1, sticky="nsew", padx=12, pady=6)
        issue_text.insert(tk.END, str(initial.get("issue", "")))
        window.rowconfigure(row, weight=1)
        row += 1

        ttk.Label(window, text="조치내용").grid(row=row, column=0, sticky="nw", padx=12, pady=6)
        action_text = tk.Text(window, height=6, wrap="word", font=("맑은 고딕", 10))
        action_text.grid(row=row, column=1, sticky="nsew", padx=12, pady=6)
        action_text.insert(tk.END, str(initial.get("action", "")))
        window.rowconfigure(row, weight=1)
        row += 1

        checks = ttk.Frame(window)
        checks.grid(row=row, column=1, sticky="w", padx=12, pady=6)
        ttk.Checkbutton(checks, text="APC", variable=apc_var).pack(side=tk.LEFT, padx=(0, 12))
        ttk.Checkbutton(checks, text="PC filter", variable=pc_filter_var).pack(side=tk.LEFT, padx=(0, 12))
        ttk.Checkbutton(checks, text="UTMP", variable=utmp_var).pack(side=tk.LEFT)
        row += 1

        def save() -> None:
            values = {
                "sequence": sequence_var.get().strip(),
                "date": date_var.get().strip(),
                "department": department_var.get().strip(),
                "user": user_var.get().strip(),
                "issue": issue_text.get("1.0", tk.END).strip(),
                "action": action_text.get("1.0", tk.END).strip(),
                "apc": "O" if apc_var.get() else "",
                "pc_filter": "O" if pc_filter_var.get() else "",
                "utmp": "O" if utmp_var.get() else "",
            }
            if month_values is not None and not month_var.get():
                messagebox.showwarning(APP_TITLE, "대상 월을 선택하세요.", parent=window)
                return
            if not values["sequence"]:
                messagebox.showwarning(APP_TITLE, "순번을 입력하세요.", parent=window)
                return
            if not values["sequence"].isdigit():
                messagebox.showwarning(APP_TITLE, "순번은 숫자로 입력하세요.", parent=window)
                return
            if month_values is not None and not values["date"]:
                messagebox.showwarning(APP_TITLE, "날짜를 입력하세요.", parent=window)
                return
            if not values["department"]:
                messagebox.showwarning(APP_TITLE, "부서를 입력하세요.", parent=window)
                return
            if not values["issue"] and not values["action"]:
                messagebox.showwarning(APP_TITLE, "장애내용 또는 조치내용을 입력하세요.", parent=window)
                return
            try:
                saved_path, success_message = on_save(values, month_var.get())
            except PermissionError:
                messagebox.showerror(
                    APP_TITLE,
                    "Excel 파일을 저장할 수 없습니다.\n\n원본 파일이 열려 있으면 닫은 뒤 다시 시도하세요.",
                    parent=window,
                )
                return
            except Exception as exc:
                messagebox.showerror(APP_TITLE, f"저장 실패\n\n{exc}", parent=window)
                return
            window.destroy()
            self.root.after(10, lambda: self._after_excel_write(saved_path, success_message))

        buttons = ttk.Frame(window)
        buttons.grid(row=row, column=0, columnspan=2, sticky="e", padx=12, pady=(8, 12))
        ttk.Button(buttons, text="저장", command=save).pack(side=tk.LEFT)
        ttk.Button(buttons, text="취소", command=window.destroy).pack(side=tk.LEFT, padx=(6, 0))

    # 새 내역 추가 시 대상 월 파일의 마지막 순번을 기준으로 다음 번호를 제안한다.
    def _get_available_source_months(self) -> list[str]:
        folder = Path(self.folder_var.get()).expanduser()
        if not folder.exists():
            return []
        months: set[str] = set()
        for path in folder.rglob("*.xlsx"):
            if get_excel_skip_reason(path):
                continue
            year, month = parse_year_month_from_path(path)
            if year and month:
                months.add(f"{year}-{month:02d}")
        return sorted(months)

    def _find_workbook_for_month(self, month_text: str) -> Optional[Path]:
        try:
            year_text, month_part = month_text.split("-", 1)
            target = (int(year_text), int(month_part))
        except ValueError:
            return None
        folder = Path(self.folder_var.get()).expanduser()
        matches: list[Path] = []
        for path in sorted(folder.rglob("*.xlsx")):
            if get_excel_skip_reason(path):
                continue
            if parse_year_month_from_path(path) == target:
                matches.append(path)
        return matches[0] if matches else None

    def _get_next_sequence_for_month(self, month_text: str) -> str:
        path = self._find_workbook_for_month(month_text)
        if path is None:
            return ""
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = self._find_maintenance_sheet(workbook)
            return str(self._next_sequence_number(sheet))
        finally:
            workbook.close()

    def _find_maintenance_sheet(self, workbook: object) -> object:
        for sheet in workbook.worksheets:
            if looks_like_maintenance_sheet(sheet):
                return sheet
        raise RuntimeError("유지보수 양식 시트를 찾을 수 없습니다.")

    def _find_next_insert_row(self, sheet: object) -> int:
        last_row = 4
        for row in range(max(sheet.max_row, 5), 4, -1):
            if any(safe_cell(sheet.cell(row, col).value) for col in range(1, 10)):
                last_row = row
                break
        return max(last_row + 1, 5)

    def _next_sequence_number(self, sheet: object) -> int:
        numbers: list[int] = []
        for row in range(5, sheet.max_row + 1):
            value = sheet.cell(row, 1).value
            try:
                numbers.append(int(value))
            except (TypeError, ValueError):
                continue
        return max(numbers, default=0) + 1

    def _copy_row_style(self, sheet: object, source_row: int, target_row: int) -> None:
        if source_row < 5:
            return
        # 새 행이 기존 유지보수 양식과 같은 모양을 유지하도록 바로 윗행의 서식을 복사한다.
        for col in range(1, 10):
            source = sheet.cell(source_row, col)
            target = sheet.cell(target_row, col)
            if source.has_style:
                target._style = copy(source._style)
            target.number_format = source.number_format
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.protection = copy(source.protection)
        sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height

    def _write_case_values(self, sheet: object, row: int, values: dict[str, str], sequence: Optional[int] = None) -> None:
        if sequence is not None and not values.get("sequence"):
            values["sequence"] = str(sequence)
        columns = {
            1: "sequence",
            2: "date",
            3: "department",
            4: "user",
            5: "issue",
            6: "action",
            7: "apc",
            8: "pc_filter",
            9: "utmp",
        }
        for col, key in columns.items():
            value = values.get(key, "")
            if key == "sequence" and value:
                sheet.cell(row, col).value = int(value)
            else:
                sheet.cell(row, col).value = value if value else None

    def _save_new_case(self, month_text: str, values: dict[str, str]) -> tuple[Path, str]:
        path = self._find_workbook_for_month(month_text)
        if path is None:
            raise RuntimeError(f"{month_text} 유지보수 엑셀 파일을 찾을 수 없습니다.")

        workbook = load_workbook(path)
        try:
            sheet = self._find_maintenance_sheet(workbook)
            insert_row = self._find_next_insert_row(sheet)
            self._copy_row_style(sheet, insert_row - 1, insert_row)
            self._write_case_values(sheet, insert_row, values)
            workbook.save(path)
        finally:
            workbook.close()
        return path, "새 내역을 추가했습니다."

    def _save_existing_case(self, case: MaintenanceCase, values: dict[str, str]) -> tuple[Path, str]:
        path = self._resolve_case_path(case)
        if path is None:
            raise RuntimeError(f"원본 파일을 찾을 수 없습니다: {case.source_file}")

        workbook = load_workbook(path)
        try:
            if case.source_sheet not in workbook.sheetnames:
                raise RuntimeError(f"원본 시트를 찾을 수 없습니다: {case.source_sheet}")
            sheet = workbook[case.source_sheet]
            self._write_case_values(sheet, case.row_num, values)
            workbook.save(path)
        finally:
            workbook.close()

        case.sequence_number = values["sequence"]
        case.date_text = values["date"]
        case.department = values["department"]
        case.user = values["user"]
        case.issue_text = values["issue"]
        case.action_text = values["action"]
        case.apc = values["apc"]
        case.pc_filter = values["pc_filter"]
        case.utmp = values["utmp"]
        case.search_text = " ".join(
            part
            for part in [case.issue_text, case.action_text, case.department, case.user, case.date_text, case.sheet_title]
            if part
        )
        self._refresh_selected_tree_row()
        self._show_selected_detail()
        return path, "선택 항목을 수정했습니다."

    def _refresh_selected_tree_row(self) -> None:
        selection = self.tree.selection()
        if not selection:
            return
        idx = self.tree.index(selection[0])
        if idx >= len(self.search_results):
            return
        result = self.search_results[idx]
        case: MaintenanceCase = result["case"]
        is_list_mode = result.get("mode") == "list"
        self.tree.item(
            selection[0],
            values=(
                result["rank"],
                "-" if is_list_mode else self._format_percent(result["score"]),
                case.year,
                case.date_text,
                case.department,
                case.user,
                self._shorten(case.issue_text, 36),
                self._shorten(case.action_text, 52),
                case.source_file,
            ),
        )

    def _after_excel_write(self, path: Path, message: str) -> None:
        folder = self.engine.loaded_folder or Path(self.folder_var.get()).expanduser()
        stale_message = "프로그램에서 Excel 파일을 수정했습니다."
        self.engine.index_stale = True
        self.engine.index_stale_message = stale_message
        self._set_index_notice(True, stale_message)
        if folder.exists():
            self._last_stale_warning_key = f"{folder.resolve()}|{stale_message}"
        self.status_var.set(f"{message} 인덱스를 새로 구축해야 최신 검색 결과가 반영됩니다.")
        if messagebox.askyesno(
            APP_TITLE,
            f"{message}\n\n변경 내용을 검색 결과에 반영하려면 인덱스를 새로 구축해야 합니다.\n지금 새로 구축할까요?",
        ):
            self._start_build(folder)

    def _set_busy(self, busy: bool) -> None:
        if busy:
            self.progress.start(10)
        else:
            self.progress.stop()
        state = ["disabled"] if busy else ["!disabled"]
        for child in self.root.winfo_children():
            self._set_state_recursive(child, state)
        # keep bottom bar usable
        self.progress.configure(mode="indeterminate")

    def _set_state_recursive(self, widget: tk.Widget, state: list[str]) -> None:
        try:
            if isinstance(widget, (ttk.Entry, ttk.Button, ttk.Combobox, ttk.Checkbutton, ttk.Spinbox)):
                widget.state(state)
        except Exception:
            pass
        for child in widget.winfo_children():
            self._set_state_recursive(child, state)

    def _get_month_values(self) -> list[str]:
        return sorted(
            {
                f"{record.year}-{record.month:02d}"
                for record in self.engine.records
                if record.year and record.month
            }
        )

    def _get_month_values_for_selected_years(self) -> list[str]:
        if not self.month_values:
            return []

        def parse_year(value: str) -> Optional[int]:
            if not value or value == "전체":
                return None
            try:
                return int(value)
            except ValueError:
                return None

        start_year = parse_year(self.year_from_var.get())
        end_year = parse_year(self.year_to_var.get())
        if start_year is None and end_year is None:
            return self.month_values
        if start_year is None:
            start_year = end_year
        if end_year is None:
            end_year = start_year
        if start_year is None or end_year is None:
            return self.month_values
        if start_year > end_year:
            start_year, end_year = end_year, start_year

        filtered: list[str] = []
        for value in self.month_values:
            try:
                year = int(value.split("-", 1)[0])
            except ValueError:
                continue
            if start_year <= year <= end_year:
                filtered.append(value)
        return filtered

    def _refresh_filter_combos(self) -> None:
        year_values = ["전체"] + [str(year) for year in self.year_values]
        self.year_from_combo.configure(values=year_values)
        self.year_to_combo.configure(values=year_values)
        if self.year_from_var.get() not in year_values:
            self.year_from_var.set("전체")
        if self.year_to_var.get() not in year_values:
            self.year_to_var.set("전체")
        self._normalize_filter_range("year_from", refresh_months=False)
        self._refresh_month_combos()

    def _refresh_month_combos(self) -> None:
        month_values = ["전체"] + self._get_month_values_for_selected_years()
        self.month_from_combo.configure(values=month_values)
        self.month_to_combo.configure(values=month_values)
        if self.month_from_var.get() not in month_values:
            self.month_from_var.set("전체")
        if self.month_to_var.get() not in month_values:
            self.month_to_var.set("전체")
        self._normalize_filter_range("month_from", refresh_months=False)

    def _normalize_filter_range(self, changed: str, refresh_months: bool = True) -> None:
        is_year = changed.startswith("year")
        if is_year:
            values = [str(year) for year in self.year_values]
            from_var = self.year_from_var
            to_var = self.year_to_var
            changed_from = changed == "year_from"
        else:
            values = self._get_month_values_for_selected_years()
            from_var = self.month_from_var
            to_var = self.month_to_var
            changed_from = changed == "month_from"

        def finish() -> None:
            if is_year and refresh_months:
                self._refresh_month_combos()

        if not values:
            from_var.set("전체")
            to_var.set("전체")
            finish()
            return

        start = from_var.get()
        end = to_var.get()
        if start != "전체" and start not in values:
            from_var.set("전체")
            start = "전체"
        if end != "전체" and end not in values:
            to_var.set("전체")
            end = "전체"

        if start == "전체" and end == "전체":
            finish()
            return

        if changed_from:
            if start == "전체":
                to_var.set("전체")
                finish()
                return
            if end == "전체":
                to_var.set(start)
                finish()
                return
            if values.index(start) > values.index(end):
                to_var.set(start)
        else:
            if end == "전체":
                from_var.set("전체")
                finish()
                return
            if start == "전체":
                from_var.set(end)
                finish()
                return
            if values.index(start) > values.index(end):
                from_var.set(end)
        finish()

    def _collect_filters(self) -> SearchFilters:
        def parse_year(value: str) -> Optional[int]:
            value = value.strip()
            if not value or value == "전체":
                return None
            try:
                return int(value)
            except ValueError:
                return None

        def parse_month(value: str) -> Optional[tuple[int, int]]:
            value = value.strip()
            if not value or value == "전체":
                return None
            try:
                year_text, month_text = value.split("-", 1)
                return int(year_text), int(month_text)
            except ValueError:
                return None

        return SearchFilters(
            year_from=parse_year(self.year_from_var.get()),
            year_to=parse_year(self.year_to_var.get()),
            month_from=parse_month(self.month_from_var.get()),
            month_to=parse_month(self.month_to_var.get()),
            department=self.department_var.get(),
            user=self.user_var.get(),
            require_apc=self.apc_var.get(),
            require_pc_filter=self.pc_filter_var.get(),
            require_utmp=self.utmp_var.get(),
        )

    def _run_search(self) -> None:
        if not self.engine.is_ready:
            messagebox.showwarning(APP_TITLE, "먼저 인덱스를 구축하거나 불러오세요.")
            return
        query = self.query_var.get().strip()
        if not query or query == QUERY_HINT:
            messagebox.showwarning(APP_TITLE, "검색어를 입력하세요.")
            return

        try:
            filters = self._collect_filters()
            results = self.engine.search(query, filters, top_n=self.top_n_var.get())
            self._show_results(results, empty_message="검색 결과가 없습니다. 검색어를 줄이거나 필터를 해제해 보세요.")
            self.status_var.set(f"검색 완료: {len(results)}건")
        except Exception as exc:
            messagebox.showerror(APP_TITLE, f"검색 실패\n\n{exc}")

    def _show_all_cases(self) -> None:
        if not self.engine.is_ready:
            messagebox.showwarning(APP_TITLE, "먼저 인덱스를 구축하거나 불러오세요.")
            return
        try:
            filters = self._collect_filters()
            results = self.engine.list_cases(filters)
            self._show_results(results, empty_message="현재 필터에 해당하는 사례가 없습니다.")
            self.status_var.set(f"전체 사례 표시: {len(results)}건")
        except Exception as exc:
            messagebox.showerror(APP_TITLE, f"전체 사례 표시 실패\n\n{exc}")

    def _show_results(self, results: list[dict[str, object]], empty_message: str = "검색 결과가 없습니다.") -> None:
        self.search_results = results
        self.tree.delete(*self.tree.get_children())
        self._clear_detail_display()
        if not results:
            self._set_detail_empty_message(empty_message)
            return

        for result in results:
            case: MaintenanceCase = result["case"]
            is_list_mode = result.get("mode") == "list"
            self.tree.insert(
                "",
                tk.END,
                values=(
                    result["rank"],
                    "-" if is_list_mode else self._format_percent(result["score"]),
                    case.year,
                    case.date_text,
                    case.department,
                    case.user,
                    self._shorten(case.issue_text, 36),
                    self._shorten(case.action_text, 52),
                    case.source_file,
                ),
            )
        self.tree.selection_set(self.tree.get_children()[0])
        self.tree.focus(self.tree.get_children()[0])
        self._show_selected_detail()

    def _shorten(self, text: str, limit: int) -> str:
        return text if len(text) <= limit else text[: limit - 1] + "…"

    def _clear_detail_display(self) -> None:
        self.detail_rank_var.set("")
        self.detail_score_var.set("")
        self.detail_date_var.set("")
        self.detail_person_var.set("")
        self.detail_flags_var.set("")
        self.detail_source_var.set("")
        self._set_text_value(self.detail_issue_text, "")
        self._set_text_value(self.detail_action_text, "")

    def _set_detail_empty_message(self, message: str) -> None:
        self.detail_rank_var.set(message)
        self.detail_score_var.set("")
        self.detail_date_var.set("")
        self.detail_person_var.set("")
        self.detail_flags_var.set("")
        self.detail_source_var.set("")
        self._set_text_value(self.detail_issue_text, "")
        self._set_text_value(self.detail_action_text, "")

    def _set_text_value(self, widget: tk.Text, value: str) -> None:
        widget.configure(state=tk.NORMAL)
        widget.delete("1.0", tk.END)
        widget.insert(tk.END, value)
        widget.configure(state=tk.DISABLED)

    def _format_case_date(self, case: MaintenanceCase) -> str:
        text = normalize_text(case.date_text)
        if text:
            match = re.search(r"(20\d{2})[.\-/년\s]+(\d{1,2})[.\-/월\s]+(\d{1,2})", text)
            if match:
                return f"{int(match.group(1))}년 {int(match.group(2))}월 {int(match.group(3))}일"

            match = re.search(r"(\d{1,2})\s*월\s*(\d{1,2})\s*일", text)
            if match and case.year:
                return f"{case.year}년 {int(match.group(1))}월 {int(match.group(2))}일"

            match = re.search(r"(\d{1,2})[.\-/](\d{1,2})", text)
            if match and case.year:
                return f"{case.year}년 {int(match.group(1))}월 {int(match.group(2))}일"

            return text
        if case.year and case.month:
            return f"{case.year}년 {case.month}월"
        if case.year:
            return f"{case.year}년"
        return "-"

    def _show_selected_detail(self, event: object | None = None) -> None:
        selection = self.tree.selection()
        if not selection:
            return
        idx = self.tree.index(selection[0])
        if idx >= len(self.search_results):
            return
        result = self.search_results[idx]
        case: MaintenanceCase = result["case"]
        is_list_mode = result.get("mode") == "list"
        self.detail_rank_var.set(f"순위 {result['rank']}")
        if is_list_mode:
            self.detail_score_var.set("관련도 -")
        else:
            self.detail_score_var.set(f"관련도 {self._format_percent(result['score'])}")
        self.detail_date_var.set(f"일자 {self._format_case_date(case)}")
        self.detail_person_var.set(f"부서 {case.department or '-'} · 사용자 {case.user or '-'}")
        self.detail_flags_var.set(
            f"APC {case.apc or '-'} · PC filter {case.pc_filter or '-'} · UTMP {case.utmp or '-'}"
        )
        self.detail_source_var.set(
            "원본 "
            f"{case.source_file} · {case.source_sheet} · {case.row_num}행 · 순번 {case.sequence_number or '-'}"
            + (f" · {case.sheet_title}" if case.sheet_title else "")
        )
        self._set_text_value(self.detail_issue_text, case.issue_text or "-")
        self._set_text_value(self.detail_action_text, case.action_text or "-")

    @staticmethod
    def _format_percent(value: object) -> str:
        number = float(value)
        return f"{max(0.0, min(number, 1.0)) * 100:.1f}%"

    def _export_results(self) -> None:
        if not self.search_results:
            messagebox.showwarning(APP_TITLE, "먼저 검색을 실행하세요.")
            return

        output_path = filedialog.asksaveasfilename(
            title="검색 결과 엑셀 저장",
            defaultextension=".xlsx",
            filetypes=[("Excel 통합 문서", "*.xlsx")],
            initialfile="유지보수_유사사례_검색결과.xlsx",
        )
        if not output_path:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "유사사례 검색결과"
            headers = [
                "순위",
                "관련도",
                "원본순번",
                "연도",
                "날짜",
                "부서",
                "사용자",
                "장애내용",
                "조치내용",
                "APC",
                "PC filter",
                "UTMP",
                "원본파일",
                "시트",
                "행번호",
            ]
            worksheet.append(headers)

            for result in self.search_results:
                case: MaintenanceCase = result["case"]
                is_list_mode = result.get("mode") == "list"
                worksheet.append(
                    [
                        result["rank"],
                        "-" if is_list_mode else self._format_percent(result["score"]),
                        case.sequence_number,
                        case.year,
                        case.date_text,
                        case.department,
                        case.user,
                        case.issue_text,
                        case.action_text,
                        case.apc,
                        case.pc_filter,
                        case.utmp,
                        case.source_file,
                        case.source_sheet,
                        case.row_num,
                    ]
                )

            header_fill = PatternFill("solid", fgColor="D9EAF7")
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            widths = [8, 12, 10, 8, 12, 18, 14, 42, 55, 9, 11, 9, 24, 16, 10]
            for index, width in enumerate(widths, start=1):
                worksheet.column_dimensions[get_column_letter(index)].width = width
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            workbook.save(output_path)
            self.status_var.set(f"엑셀 저장 완료: {output_path}")
            messagebox.showinfo(APP_TITLE, "검색 결과를 엑셀 파일로 저장했습니다.")
        except Exception as exc:
            messagebox.showerror(APP_TITLE, f"엑셀 저장 실패\n\n{exc}")


def main() -> None:
    root = tk.Tk()
    try:
        style = ttk.Style()
        theme_names = style.theme_names()
        if sys.platform == "win32" and "vista" in theme_names:
            style.theme_use("vista")
        elif sys.platform == "darwin" and "aqua" in theme_names:
            style.theme_use("aqua")
        elif "clam" in theme_names:
            style.theme_use("clam")
    except Exception:
        pass
    MaintenanceSearchApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
